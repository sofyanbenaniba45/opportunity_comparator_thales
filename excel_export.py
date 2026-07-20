"""
Excel report generation.
"""
import datetime
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Colors
FILL_NEW           = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
FILL_MODIFIED_CELL = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")  # Orange
FILL_HEADER        = PatternFill(start_color="244061", end_color="244061", fill_type="solid")  # Dark blue

FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)

THIN = Side(border_style="thin", color="AAAAAA")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _normalize_key(val) -> str:
    return str(val).strip().lower()


def _gbu_label(v) -> str:
    s = str(v).strip()
    return "(No GBU/BL)" if s in ("", "nan", "NaT", "None", "<NA>") else s


def _ou_label(v) -> str:
    s = str(v).strip()
    return "(No OU/LE)" if s in ("", "nan", "NaT", "None", "<NA>") else s


def _write_sheet(ws, columns: list, df_gbu: pd.DataFrame, diff_map: dict):
    """Write header + data rows to a worksheet."""
    # Header row
    ws.append(columns)
    for col_idx, _ in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN

    # Data rows
    for _, row in df_gbu.iterrows():
        status = row.get("__status__", "unchanged")
        norm_key = _normalize_key(row.get("Opportunity Name", ""))
        changed_cols = diff_map.get(norm_key, set())

        row_vals = [row.get(c, "") for c in columns]
        ws.append(row_vals)

        row_idx = ws.max_row
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = BORDER_THIN
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            if isinstance(cell.value, (datetime.date, datetime.datetime, pd.Timestamp)):
                cell.number_format = "yyyy-mm-dd"

            if status == "new":
                cell.fill = FILL_NEW
            elif status == "modified" and col_name in changed_cols:
                cell.fill = FILL_MODIFIED_CELL

    # Auto column width
    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col_name))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    cell_len = len(str(cell.value)) if cell.value else 0
                    max_len = max(max_len, min(cell_len, 40))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = max(max_len + 2, 12)

    ws.freeze_panes = "A2"


def generate_excel(result: dict, gbu_col: str) -> bytes:
    """
    Generate the Excel report in memory and return bytes.

    Sheet 1 : All Opportunities (all GBU/BL combined)
    Sheet 2+ : one sheet per GBU/BL value
    """
    columns  = result["columns"]
    diff_map = result["diff_map"]

    # Rows to export: new + modified + unchanged (no deleted)
    frames = []
    for status in ("new", "modified", "unchanged"):
        df = result[status]
        if not df.empty:
            frames.append(df)

    wb = Workbook()
    wb.remove(wb.active)

    if not frames:
        ws = wb.create_sheet(title="Report")
        ws.append(["No data to export."])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    df_all = pd.concat(frames, ignore_index=True).copy()

    # Build GBU label column
    if gbu_col and gbu_col in df_all.columns:
        df_all["__gbu__"] = [_gbu_label(v) for v in df_all[gbu_col]]
    else:
        df_all["__gbu__"] = "All"

    gbu_values = sorted(df_all["__gbu__"].unique())

    # ── Sheet 1: All Opportunities ──
    ws_all = wb.create_sheet(title="All Opportunities")
    _write_sheet(ws_all, columns, df_all, diff_map)

    # ── One sheet per GBU/BL ──
    groups = df_all.groupby("__gbu__")
    for gbu in gbu_values:
        df_gbu = groups.get_group(gbu)
        sheet_name = str(gbu)[:31]
        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, columns, df_gbu, diff_map)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_excel_by_ou(result: dict, ou_col: str, gbu_col: str) -> bytes:
    """
    Generate an Excel report grouped by Operating Unit / Legal Entity.

    Sheet 1 : All Opportunities
    Sheet 2+: one sheet per OU/LE (all GBU/BL rows combined within that OU/LE)
    """
    columns  = result["columns"]
    diff_map = result["diff_map"]

    frames = []
    for status in ("new", "modified", "unchanged"):
        df = result[status]
        if not df.empty:
            frames.append(df)

    wb = Workbook()
    wb.remove(wb.active)

    if not frames:
        ws = wb.create_sheet(title="Report")
        ws.append(["No data to export."])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    df_all = pd.concat(frames, ignore_index=True).copy()

    if ou_col and ou_col in df_all.columns:
        df_all["__ou__"] = [_ou_label(v) for v in df_all[ou_col]]
    else:
        df_all["__ou__"] = "All"

    ou_values = sorted(df_all["__ou__"].unique())

    # Sheet 1: all opportunities
    ws_all = wb.create_sheet(title="All Opportunities")
    _write_sheet(ws_all, columns, df_all, diff_map)

    # One sheet per OU/LE
    groups = df_all.groupby("__ou__")
    for ou in ou_values:
        df_ou = groups.get_group(ou)
        sheet_name = str(ou)[:31]
        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, columns, df_ou, diff_map)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
